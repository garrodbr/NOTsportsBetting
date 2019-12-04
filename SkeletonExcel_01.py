# ------
# Goals
#   - Scrape website to find games
#   - Populate excel sheet with games
#   - Populate excel sheet with formulas
#   - This is purely to develop the skeleton document, not to update the scores
# -------


# Get all bowl games
#   Scrape URL for bowl games
#   Get bowl game name, team, and date
#   INPUT: URL
#   OUTPUT: Bowl Name, Date, Home Team, Away Team
#       output could be a class, list of lists, dict, etc.
#       should sort by date
#       List of lists would be very simple, class would be inclusive

import requests
from bs4 import BeautifulSoup
# -------------------------------------------------------------------------------------------------------------------
# Basic download of page
# -------------------------------------------------------------------------------------------------------------------

url = "https://www.ncaa.com/scoreboard/football/fbs/2019/14"
page = requests.get(url)
soup = BeautifulSoup(page.content, 'html.parser')

# -------------------------------------------------------------------------------------------------------------------
# Get all containers, they contain the date
# -------------------------------------------------------------------------------------------------------------------

container = soup.find_all(class_="gamePod_content-division")
# date = container[0].find('h6')

# -------------------------------------------------------------------------------------------------------------------
# Get all the games
# -------------------------------------------------------------------------------------------------------------------

gamesList = []
for pod in container:
    date = pod.find('h6').get_text()
    teamNames = [team.get_text() for team in pod.find_all(class_="gamePod-game-team-name")]
    # teamScores = [score.get_text() for score in pod.find_all(class_="gamePod-game-team-score")]
    for index in range(0,len(teamNames), 2):
        homeTeam = teamNames[index]
        awayTeam = teamNames[index + 1]
        gameIndex = [date, 'BOWL NAME' + str(index), homeTeam, awayTeam]
        gamesList.append(gameIndex)

# -------------------------------------------------------------------------------------------------------------------
# Create the worksheet
# -------------------------------------------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.styles import PatternFill

players = ["Brad", "Burton", "Chris", "Martin"]
playerEntry = []
for user in players:
    playerEntry += [user + 'Home', user + 'Away', user + 'Score']

wb = Workbook()
destFile = "sampleSkeleton.xlsx"

# Change the Cover sheet name
ws1 = wb.active
ws1.title = "CoverPage"

# Create a new sheet for the bowl games
wb.create_sheet("BowlGames")
wb.active = wb["BowlGames"]

# Create the Header
ws = wb.active
appendHeader = ['DATE', 'BOWL NAME', 'HOME TEAM', 'AWAY TEAM'] + playerEntry + ['ACTUAL HOME', 'ACTUAL AWAY']
ws.append(appendHeader)

# Fill in the games
for game in gamesList:
    ws.append(game)

# Fill in the formula
# =IF(ISBLANK($Q2), , SUM(IF(OR(AND($Q2>$R2, E2>F2), AND($R2>$Q2, F2>E2)), 50, 0), E2-$Q2, F2-$R2)) sample for 4 users
# Q = Home Team Actual
# R = Away Team Actual
# E = Home Team User
# F = Away Team User
# 2 = Row Number

scoreFormula = "=IF(ISBLANK(${}{}), , SUM(IF(OR(AND(${}{}>${}{}, {}{}>{}{}), AND(${}{}>${}{}, {}{}>{}{})), 50, 0), " \
               "50-ABS({}{}-${}{})-ABS({}{}-${}{})))"  # order is Q, Q, R, E, F, R, Q, F, E, E, Q, F, R
headerLength = 0 # This should get overwritten
for rowi, rows in enumerate(ws.iter_rows()):

    if rowi == 0:
        headerLength = len(rows)
        continue # Skip the header

    rowi += 1

    for celli, cell in enumerate(rows):

        if celli in [6, 9, 12, 15]: # still could update to be auto based on number of entries
            cell.fill = PatternFill(fgColor="deb137", fill_type = "solid") # Format color
            changeCell = cell
            cellQ = chr(64+headerLength-1)
            cellR = chr(64+headerLength)
            cellE = chr(64 + celli-1)
            cellF = chr(64 + celli)

            cellFormula = scoreFormula.format(
                cellQ, rowi, cellQ, rowi, cellR, rowi, cellE, rowi, cellF, rowi, cellR, rowi, cellQ, rowi, cellF, rowi,
                cellE, rowi, cellE, rowi, cellQ, rowi, cellF, rowi, cellR, rowi)
            # order is Q, Q, R, E, F, R, Q, F, E, E, Q, F, R

            changeCell.value = cellFormula

wb.save(filename=destFile)

