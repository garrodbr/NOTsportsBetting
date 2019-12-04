# ------
# Goals
#   - Scrape website to find scores
#   - Update excel sheet with scores
# -------

# Get all bowl games scores
#   Scrape URL for bowl games scores
#   Match score with teamNames
#   INPUT: URL
#   OUTPUT: scores in excel

import requests
from bs4 import BeautifulSoup
# -------------------------------------------------------------------------------------------------------------------
# Basic download of page
# -------------------------------------------------------------------------------------------------------------------

url = "https://www.ncaa.com/scoreboard/football/fbs/2019/14"
page = requests.get(url)
soup = BeautifulSoup(page.content, 'html.parser')

# -------------------------------------------------------------------------------------------------------------------
# Get all the games
# -------------------------------------------------------------------------------------------------------------------

teamNames = [team.get_text() for team in soup.find_all(class_="gamePod-game-team-name")]
teamScores = [score.get_text() for score in soup.find_all(class_="gamePod-game-team-score")]
gameIndex = []
for index in range(0,len(teamNames), 2):
    homeTeam = teamNames[index]
    awayTeam = teamNames[index + 1]
    homeScore = teamScores[index]
    awayScore = teamScores[index+1]
    gameIndex.append([homeTeam, homeScore, awayTeam, awayScore])# COULD CREATE CLASS HERE

# -------------------------------------------------------------------------------------------------------------------
# Update the worksheet
# -------------------------------------------------------------------------------------------------------------------
import openpyxl

fileName = "sampleSkeleton.xlsx"

# To open the workbook
wb = openpyxl.load_workbook(fileName)

ws = wb['BowlGames']

headerLength = 0 # This should get overwritten
for rowi, rows in enumerate(ws.iter_rows()):

    if rowi == 0:
        headerLength = len(rows)
        continue # Skip the header

    rowi += 1 # correct the index

    homeTeam = rows[2].value
    awayTeam = rows[3].value

    for game in gameIndex:
        if homeTeam in game:
            homeScore = game[1]
            awayScore = game[3]
            ws.cell(row=rowi, column=headerLength).value = awayScore
            ws.cell(row=rowi, column=headerLength-1).value = homeScore

wb.save(filename=fileName)
