# ------
# Goals
#   - Scrape website to find games
#   - Populate excel sheet with games
#   - Populate excel sheet with formulas
#   - This is purely to develop the skeleton document, not to update the scores
# -------


# Get all bowl games
#   Scrape URL for bowl games
#   Get bowl game name, team, and date
#   INPUT: URL
#   OUTPUT: Bowl Name, Date, Home Team, Away Team
#       output could be a class, list of lists, dict, etc.
#       should sort by date
#       List of lists would be very simple, class would be inclusive

import requests
from bs4 import BeautifulSoup
# -------------------------------------------------------------------------------------------------------------------
# Basic download of page
# -------------------------------------------------------------------------------------------------------------------

# url = "https://www.ncaa.com/news/football/article/2019-12-07/2019-20-college-football-bowl-schedule-dates-times-tv-channels"
url = "https://www.cbssports.com/college-football/news/2019-bowl-games-schedule-college-football-playoff-teams-ncaa-bowl-kickoff-times/"
page = requests.get(url)
soup = BeautifulSoup(page.content, 'html.parser')

# -------------------------------------------------------------------------------------------------------------------
# Get all containers, they contain the date
# -------------------------------------------------------------------------------------------------------------------

tables = soup.find_all('table') # The games are in multiple tables

# -------------------------------------------------------------------------------------------------------------------
# Get all the games
# -------------------------------------------------------------------------------------------------------------------
import re
gamesList = []
for table in tables:
    for rows in table.find_all('tr'):
        cells = rows.find_all('td')
        if len(cells)>0:
            date = cells[0].get_text().strip()
            location = cells[1]
            locSplit = str(location).split("<br/><em>")
            bowlName = re.sub("<p>|<td>|</em>|</td>|</p>", "", locSplit[0]).strip()
            bowlLocation = re.sub("<p>|<td>|</em>|</td>|</p>", "", locSplit[1]).strip()
            time = cells[2].get_text().strip()
            teams = cells[3].get_text().strip().split(' vs. ')
            if len(teams) < 2:
                teams = ["TEAM1", "TEAM2"]
            entry = [date, bowlName, bowlLocation, teams[0], '', teams[1], ''] # added the extra entries for excel reasons
            gamesList.append(entry)

# -------------------------------------------------------------------------------------------------------------------
# Create the worksheet
# -------------------------------------------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.styles import PatternFill

players = ["Brad", "Burton", "Chris", "Martin", "P5", "P6", "P7", "P8"]
playerEntry = []
for user in players:
    playerEntry += [user + ' Home', user + ' Away', user + ' Score']

wb = Workbook()
destFile = "sampleSkeleton2.xlsx"

# Change the Cover sheet name
ws1 = wb.active
ws1.title = "CoverPage"

# Create a new sheet for the bowl games
wb.create_sheet("BowlGames")
wb.active = wb["BowlGames"]

# Create the Header
ws = wb.active
appendHeader = ['DATE', 'BOWL NAME', 'BOWL LOCATION', 'HOME TEAM', 'HOME SCORE', 'AWAY TEAM', 'AWAY SCORE'] + playerEntry
headerLength = len(appendHeader)
playerCellList = [*range(9, headerLength, 3)]
print(playerCellList)
ws.append(appendHeader)

# Fill in the games
for game in gamesList:
    ws.append(game)

# Fill in the formula
# =IF(ISBLANK($E2), , SUM(IF(OR(AND($Q2>$R2, E2>F2), AND($R2>$Q2, F2>E2)), 50, 0), E2-$Q2, F2-$R2)) sample for 4 users
# Q = Home Team Actual
# R = Away Team Actual
# E = Home Team User
# F = Away Team User
# 2 = Row Number

scoreFormula = "=IF(ISBLANK(${}{}), , SUM(IF(OR(AND(${}{}>${}{}, {}{}>{}{}), AND(${}{}>${}{}, {}{}>{}{})), 50, 0), " \
               "50-ABS({}{}-${}{})-ABS({}{}-${}{})))"  # order is Q, Q, R, E, F, R, Q, F, E, E, Q, F, R

for rowi, rows in enumerate(ws.iter_rows()):

    if rowi == 0:
        continue # Skip the header

    rowi += 1

    for celli, cell in enumerate(rows):

        if celli in playerCellList: # still could update to be auto based on number of entries
            cell.fill = PatternFill(fgColor="deb137", fill_type = "solid") # Format color
            changeCell = cell
            cellQ = "E"
            cellR = "G"
            cellE = chr(64 + celli-1)
            cellF = chr(64 + celli)
            #
            cellFormula = scoreFormula.format(
                cellQ, rowi, cellQ, rowi, cellR, rowi, cellE, rowi, cellF, rowi, cellR, rowi, cellQ, rowi, cellF, rowi,
                cellE, rowi, cellE, rowi, cellQ, rowi, cellF, rowi, cellR, rowi)
            # order is Q, Q, R, E, F, R, Q, F, E, E, Q, F, R

            changeCell.value = cellFormula

wb.save(filename=destFile)

